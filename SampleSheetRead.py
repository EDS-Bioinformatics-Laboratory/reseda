import openpyxl as xlsx
import xlrd as xls
import sys
import os


def findInRow(p, s):
    '''
    In: p = search terms (list), s = string
    Out: True or False (found a match or not)
    '''
    match = False
    for pattern in p:
        if pattern in s:
            match = True

    return(match)


def readXlsx(f, p):
    # DOC: https://openpyxl.readthedocs.io/en/default/tutorial.html#loading-from-a-file
    wb = xlsx.load_workbook(f)
    sheets = wb.get_sheet_names()
    # print(sheets)
    for sheet_name in sheets:
        s = wb[sheet_name]
        # print(sheet_name, s.max_row, 'x', s.max_column)

        for row in s.rows:
            values = list()
            for cell in row:
                values.append(str(cell.value))
            myrow = ','.join(values)
            if findInRow(p, myrow) is True:
                print(f, sheet_name, myrow)


def readXls(f, p):
    # DOC: https://github.com/python-excel/tutorial/blob/master/python-excel.pdf
    wb = xls.open_workbook(filename=f)
    sheets = list()
    for s in wb.sheets():
        sheets.append(s.name)
        # print(s.name, s.nrows, 'x', s.ncols)

        for row in range(s.nrows):
            values = list()
            for col in range(s.ncols):
                values.append(str(s.cell(row, col).value))
            myrow = ','.join(values)
            if findInRow(p, myrow) is True:
                print(f, s.name, myrow)


if __name__ == '__main__':
    if len(sys.argv) < 2:
        sys.exit("Usage: " + sys.argv[0] + "MID-overview-directory-name/")
    d = sys.argv[1]
    myfiles = os.listdir(d)
    myfiles.sort()

    for myfile in myfiles:
        f = d + myfile
        # print(f)

        # Check extension: .xls .xlsx .csv and use the appropriate package to open the file
        ext = f.split(".")[-1]

        p = list()
        for i in range(25):
            p.append("HD" + str(i))
        # print("p = ", p)

        if ext == "xlsx":
            readXlsx(f, p)
        elif ext == "xls":
            readXls(f, p)
        else:
            pass
            # print("Unknown extension:", ext)
